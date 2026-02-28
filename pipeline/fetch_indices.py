#!/usr/bin/env python3
"""
Fetch preparedness indices: GHSI, INFORM, and WHO IHR SPAR.
GHSI and INFORM are downloaded as Excel files and parsed with openpyxl.
SPAR is fetched from WHO GHO API.
Outputs data/indices.json keyed by ISO3.
"""

import csv
import io
import json
import time
from pathlib import Path

import openpyxl
import requests

from country_mapping import NAME_TO_ISO3, get_iso3

DATA_DIR = Path(__file__).parent.parent / "data"
SOURCE_DIR = Path(__file__).parent / "source_data"
OUTPUT_FILE = DATA_DIR / "indices.json"

SOURCE_DIR.mkdir(exist_ok=True)

# --- GHSI ---

GHSI_URL = "https://www.ghsindex.org/wp-content/uploads/2022/04/2021-GHS-Index-April-2022.csv"
GHSI_FILE = SOURCE_DIR / "ghsi_2021.csv"

GHSI_CATEGORIES = [
    "Prevention",
    "Detection & Reporting",
    "Rapid Response",
    "Health System",
    "Compliance",
    "Risk Environment",
]


def download_file(url: str, filepath: Path) -> bool:
    """Download a file if not already cached."""
    if filepath.exists():
        print(f"  Using cached: {filepath.name}")
        return True
    print(f"  Downloading: {url}")
    try:
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        filepath.write_bytes(resp.content)
        print(f"  Saved: {filepath.name} ({len(resp.content) // 1024} KB)")
        return True
    except Exception as e:
        print(f"  Failed to download: {e}")
        return False


def parse_ghsi() -> dict[str, dict]:
    """Parse GHSI 2021 CSV file."""
    if not download_file(GHSI_URL, GHSI_FILE):
        return {}

    print("Parsing GHSI data...")
    text = GHSI_FILE.read_text(encoding="utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    fieldnames = reader.fieldnames or []
    print(f"  CSV columns: {fieldnames[:15]}")

    # Find column names — match only top-level columns, not sub-indicators
    country_col = None
    overall_col = None
    cat_cols: dict[str, str] = {}

    for f in fieldnames:
        fl = f.lower().strip()
        # Match exact "country" or "country name" column, not sub-indicators containing "country"
        if fl in ("country", "country name", "name"):
            country_col = f
        elif fl == "overall score" or fl == "overall":
            overall_col = f
        else:
            # Match top-level category columns (start with digit+paren, e.g. "1) PREVENTION...")
            if not (fl.startswith(("1)", "2)", "3)", "4)", "5)", "6)")) and ")" in fl[:3]):
                continue
            for cat in GHSI_CATEGORIES:
                if cat.lower().split()[0] in fl:
                    cat_cols[cat] = f
                    break

    if not country_col or not overall_col:
        print(f"  Could not find country/overall columns in CSV headers")
        return {}

    results = {}
    for row in reader:
        country_name = row.get(country_col, "").strip()
        if not country_name:
            continue
        iso3 = get_iso3(country_name)
        if not iso3:
            continue
        try:
            overall = float(row[overall_col])
        except (TypeError, ValueError, KeyError):
            continue

        categories = {}
        for cat_name, col_name in cat_cols.items():
            try:
                categories[cat_name] = round(float(row[col_name]), 1)
            except (TypeError, ValueError, KeyError):
                pass

        results[iso3] = {
            "iso3": iso3,
            "indexName": "GHSI",
            "score": round(overall, 1),
            "year": 2021,
            "categories": categories if categories else None,
        }

    print(f"  Parsed {len(results)} countries from GHSI")
    return results


# --- INFORM ---

INFORM_URL = "https://drmkc.jrc.ec.europa.eu/inform-index/Portals/0/InfoRM/2025/INFORM_Risk_Mid_2025_v071.xlsx"
INFORM_FILE = SOURCE_DIR / "inform_2025.xlsx"

INFORM_DIMENSIONS = {
    "Hazard & Exposure": ["hazard", "ha."],
    "Vulnerability": ["vulnerability", "vu."],
    "Lack of Coping Capacity": ["coping", "cc."],
}


def parse_inform() -> dict[str, dict]:
    """Parse INFORM Risk Index Excel file."""
    if not download_file(INFORM_URL, INFORM_FILE):
        return {}

    print("Parsing INFORM data...")
    wb = openpyxl.load_workbook(INFORM_FILE, read_only=True, data_only=True)

    # Find the main scores sheet
    target_sheet = None
    for name in wb.sheetnames:
        nl = name.lower()
        if "inform" in nl and ("risk" in nl or "score" in nl):
            target_sheet = name
            break
    if not target_sheet:
        # Try sheets with 'country' or just the first
        for name in wb.sheetnames:
            if "country" in name.lower():
                target_sheet = name
                break
    if not target_sheet:
        target_sheet = wb.sheetnames[0]

    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("  No data found in INFORM sheet")
        return {}

    # Find header row
    header_row = None
    for i, row in enumerate(rows[:10]):
        row_str = " ".join(str(c).lower() for c in row if c)
        if "iso3" in row_str or ("country" in row_str and "inform" in row_str):
            header_row = i
            break
    if header_row is None:
        header_row = 0

    headers = [str(c).strip() if c else "" for c in rows[header_row]]

    # Find columns
    iso3_col = None
    inform_col = None
    dim_cols = {}

    for i, h in enumerate(headers):
        hl = h.lower()
        if hl in ("iso3", "iso_3", "iso"):
            iso3_col = i
        elif "inform" in hl and "risk" in hl:
            inform_col = i
        elif inform_col is None and "inform" in hl:
            inform_col = i
        else:
            for dim_name, keywords in INFORM_DIMENSIONS.items():
                if any(kw in hl for kw in keywords):
                    dim_cols[dim_name] = i
                    break

    # If no ISO3 column, look for country name
    country_col = None
    if iso3_col is None:
        for i, h in enumerate(headers):
            if "country" in h.lower() or "name" in h.lower():
                country_col = i
                break

    if (iso3_col is None and country_col is None) or inform_col is None:
        print(f"  Could not find required columns. Headers: {headers[:15]}")
        return {}

    results = {}
    for row in rows[header_row + 1:]:
        if not row:
            continue

        # Get ISO3
        iso3 = None
        if iso3_col is not None and row[iso3_col]:
            val = str(row[iso3_col]).strip().upper()
            if len(val) == 3 and val.isalpha():
                iso3 = val
        if not iso3 and country_col is not None and row[country_col]:
            iso3 = get_iso3(str(row[country_col]).strip())
        if not iso3:
            continue

        try:
            overall = float(row[inform_col])
        except (TypeError, ValueError):
            continue

        categories = {}
        for dim_name, col_idx in dim_cols.items():
            try:
                categories[dim_name] = round(float(row[col_idx]) * 10, 1)  # 0-10 → 0-100
            except (TypeError, ValueError, IndexError):
                pass

        results[iso3] = {
            "iso3": iso3,
            "indexName": "INFORM",
            "score": round(overall * 10, 1),  # 0-10 → 0-100
            "year": 2025,
            "invertedScale": True,  # Higher = more risk
            "categories": categories if categories else None,
        }

    wb.close()
    print(f"  Parsed {len(results)} countries from INFORM")
    return results


# --- SPAR ---

WHO_GHO_BASE = "https://ghoapi.azureedge.net/api/"

SPAR_INDICATORS = {
    "IHRSPAR2_C01": "Legislation & financing",
    "IHRSPAR2_C02": "Policy & coordination",
    "IHRSPAR2_C03": "Zoonotic events",
    "IHRSPAR2_C04": "Food safety",
    "IHRSPAR2_C05": "Laboratory",
    "IHRSPAR2_C06": "Surveillance",
    "IHRSPAR2_C07": "Human resources",
    "IHRSPAR2_C08": "Health emergency framework",
    "IHRSPAR2_C09": "Health service provision",
    "IHRSPAR2_C10": "Risk communication",
    "IHRSPAR2_C11": "Points of entry",
    "IHRSPAR2_C12": "Chemical events",
    "IHRSPAR2_C13": "Radiation emergencies",
    "IHRSPAR2_C14": "Climate change",
    "IHRSPAR2_C15": "Communities",
}


def fetch_spar() -> dict[str, dict]:
    """Fetch WHO IHR SPAR data from GHO API."""
    print("Fetching SPAR data from WHO GHO API...")
    country_scores: dict[str, dict[str, float]] = {}

    for code, label in SPAR_INDICATORS.items():
        url = f"{WHO_GHO_BASE}{code}"
        try:
            resp = requests.get(url, timeout=30)
            resp.raise_for_status()
            data = resp.json()

            for entry in data.get("value", []):
                iso3 = entry.get("SpatialDim", "")
                if len(iso3) != 3:
                    continue
                value = entry.get("NumericValue")
                year = entry.get("TimeDim")
                if value is None or year is None:
                    continue

                # Keep most recent year
                if iso3 not in country_scores:
                    country_scores[iso3] = {"_years": {}}
                prev_year = country_scores[iso3]["_years"].get(code, 0)
                if int(year) >= prev_year:
                    country_scores[iso3][code] = float(value)
                    country_scores[iso3]["_years"][code] = int(year)

            time.sleep(0.3)
        except Exception as e:
            print(f"  Warning: Failed to fetch {code}: {e}")

    # Build results
    results = {}
    for iso3, scores in country_scores.items():
        categories = {}
        for code, label in SPAR_INDICATORS.items():
            if code in scores:
                categories[label] = round(scores[code], 1)

        if not categories:
            continue

        overall = round(sum(categories.values()) / len(categories), 1)
        results[iso3] = {
            "iso3": iso3,
            "indexName": "SPAR",
            "score": overall,
            "year": max(scores.get("_years", {}).values(), default=2024),
            "categories": categories,
        }

    print(f"  Fetched SPAR data for {len(results)} countries")
    return results


def main():
    ghsi = parse_ghsi()
    inform = parse_inform()
    spar = fetch_spar()

    # Merge into output keyed by ISO3 → list of IndexScore
    combined: dict[str, list[dict]] = {}
    for source in [ghsi, inform, spar]:
        for iso3, data in source.items():
            # Remove internal tracking fields
            clean = {k: v for k, v in data.items() if not k.startswith("_")}
            combined.setdefault(iso3, []).append(clean)

    OUTPUT_FILE.write_text(json.dumps(combined, indent=2))
    print(f"\nWrote {OUTPUT_FILE}")
    print(f"  Countries with GHSI: {len(ghsi)}")
    print(f"  Countries with INFORM: {len(inform)}")
    print(f"  Countries with SPAR: {len(spar)}")
    print(f"  Total countries in indices.json: {len(combined)}")


if __name__ == "__main__":
    main()
